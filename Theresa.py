from ast import While
from typing import Counter
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import pandas as pd
from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QLabel, QFileDialog, QTextEdit
from PyQt5 import uic
import sys
import itertools

from pyparsing import col

#Variables
Data_Label_Dict = {}
Final_Dict = {}
FieldLabel_list_DAC=['S/N','Service Type','Name of Service Provider (SP)','Region','Primary Service Group','Approved Capacity','SMM Capacity','Reported Vacancy','Full-Time','Part-Time','Total','% of Utilization','No. of PwDs admitted to the service in current month','No. of PwDs discharge from the service in current month','SGE New Referral','Accepted By SP\n(Accepted clients enrolling on a later date)','Pending\nAssessment', 'Pending\nAssessment', 'Pending Final\nAssessment Outcome','Pending Trial Enrollment', 'On Hold by\nPwD/Caregiver', 'On Hold due to \nCentre Constraint', 'On Hold by\nPwD/Caregiver','Total\n(a + b + c)','Estimated Waiting Time\n(SP Estimated Waiting Time for Enrollment)','Average Waiting Time\n(Average Waiting Time based on cases enrolled the past year)',' Median \n(by Month)','Longest Referral on Waitlist\n(Oldest Case on waitlist that does not have finalized outcome)']
FieldLabel_list_SW=['S/N','Service Type','Name of Service Provider (SP)','Region','Primary Service Group','Approved Capacity','SMM Capacity','Reported Vacancy','Full-Time','Total','% of Utilization','No. of PwDs admitted to the service in current month','No. of PwDs discharge from the service in current month','SGE New Referral','Accepted By SP (Accepted clients enrolling on a later date)','Pending Assessment','Pending Assessment','Pending Final Assessment Outcome','Pending Trial Enrollment','On Hold by PwD/Caregiver','On Hold due to Centre Constraint','On Hold by PwD/Caregiver','Total (a + b + c)','Estimated Waiting Time (SP Estimated Waiting Time for Enrollment)','Average Waiting Time (Average Waiting Time based on cases enrolled the past year)',' Median (by Month)','Longest Referral on Waitlist (Oldest Case on waitlist that does not have finalized outcome)']
FieldLabel_list_ADH=['S/N','Service Type','Name of Service Provider (SP)','Region','Primary Service Group','Approved Capacity','SMM Capacity','Reported Vacancy','Total','% of Utilization','No. of PwDs admitted to the service in current month','No. of PwDs discharge from the service in current month','SGE New Referral','Accepted By SP (Accepted clients enrolling on a later date)','Pending Assessment','Pending Assessment','Pending Final Assessment Outcome','Pending Trial Enrollment','On Hold by PwD/Caregiver','On Hold due to Centre Constraint','On Hold by PwD/Caregiver','Total (a + b + c)','Estimated Waiting Time (SP Estimated Waiting Time for Enrollment)','Average Waiting Time (Average Waiting Time based on cases enrolled the past year)',' Median (by Month)','Longest Referral on Waitlist (Oldest Case on waitlist that does not have finalized outcome)']
FieldLabel_list_Hostel=['S/N','Service Type','Name of Service Provider (SP)','Region','Primary Service Group','Approved Capacity','SMM Capacity','Reported Vacancy','Total','% of Utilization','No. of PwDs admitted to the service in current month','No. of PwDs discharge from the service in current month','SGE New Referral','Accepted By SP (Accepted clients enrolling on a later date)','Pending Assessment','Pending Assessment','Pending Final Assessment Outcome','Pending Trial Enrollment','On Hold by PwD/Caregiver','On Hold due to Centre Constraint','On Hold by PwD/Caregiver','Total (a + b + c)','Estimated Waiting Time (SP Estimated Waiting Time for Enrollment)','Average Waiting Time (Average Waiting Time based on cases enrolled the past year)',' Median (by Month)','Longest Referral on Waitlist (Oldest Case on waitlist that does not have finalized outcome)']
data_dict = {}
data_list = []
counter = 1
filtered_Dict= {}
hidden_columns=[]
hidden_list=[]

# functions for PyQt5 Window
def main_loop():
    global Data_Label_Dict, Final_Dict, data_dict, data_list, counter, filtered_Dict, hidden_columns, loop_count
    sheet_list = ['DAC','ADH','SW','Hostel']
    loop_count = 0
    for x in sheet_list :
        display(x)
        data_dict = {}
        data_list = []
        counter = 1
        filtered_Dict= {}
        hidden_columns=[]
        Data_Label_Dict = {}
        Final_Dict = {}
        loop_count+= 1

def display(PortFromSheett):
    global PortFromFile, PortToFile, PortToSheet, PortFromSheet
    global wb, ws, new_wb, new_ws, counter, data_dict
    
    PortFromSheet = PortFromSheett

    if PortFromFile != '' and PortToFile != '' and PortToSheet != '' :
        wb = load_workbook(filename= Port_From_Path)
        wb.active = wb[PortFromSheet]
        ws = wb.active

        new_wb = load_workbook(filename=Port_To_Path)
        if PortToSheet in new_wb.sheetnames:
            new_wb.active = new_wb[PortToSheet]
            new_ws = new_wb.active
        else:
            new_wb.create_sheet(PortToSheet)
            new_wb.active = new_wb[PortToSheet]
            new_ws = new_wb.active


        get_field_label() # Function to get field labels in list (data_field_list)

        if loop_count < 1:
            standard_template()

        for x in range(1,50): # Cycle through all the rows
            row_data = read_row(ws,x) # Get row_data row by row
            if  (isinstance(row_data[0], int)) == True : # check if the first column is a number if it is means the data is useful data
                #print(row_data)
                data_dict[counter] = row_data
                counter+=1
        data_dict= remove_hidden_columns(data_dict)
        #print(data_dict)
        fill_empty_cell()
        #print(data_dict)

        for x in range(1,len(data_dict)+1): # Loop through the every single row and remove empty columns
            data_dict[x] = remove_empty_columns(data_dict[x])
            #print(data_dict)
            #populate_sheet(new_ws,x,data_dict[x]) # create new sheet and populate sheet with data row by row

        #Adding Labels to Data using dictionaries
        a = list(data_dict.values())
        counter = 1
        if PortFromSheet == 'SW':
            for x in a:
                res = dict(zip(data_field_list, x)) # add labels to data
                Final_Dict[counter] = res # add labels numbering to labeled data dictionary
                counter+=1
            #print(Final_Dict)
        elif PortFromSheet == 'ADH':
            for x in a:
                res = dict(zip(data_field_list, x)) # add labels to data
                Final_Dict[counter] = res # add labels numbering to labeled data dictionary
                counter+=1
        elif PortFromSheet == 'Hostel':
            for x in a:
                res = dict(zip(data_field_list, x)) # add labels to data
                Final_Dict[counter] = res # add labels numbering to labeled data dictionary
                counter+=1
        else:
            for x in a:
                res = dict(zip(data_field_list, x)) # add labels to data
                Final_Dict[counter] = res # add labels numbering to labeled data dictionary
                counter+=1

        count = 1
        #print(Final_Dict)
        for x in list(Final_Dict.values()):
            filtered_Dict[count]={"Name of Service Provider (SP)":x["Name of Service Provider (SP)"],'Region':x['Region'],'Approved Capacity':x['Approved Capacity'],'Max Capacity':x['SMM Capacity'],'(b) Enrolment':x['Total'],'(c) Waitlist':x['Total\n(a + b + c)']}
            count+=1
        count = 0
        print(filtered_Dict)
        update_data()


        new_wb.save(Port_To_Path) # Save data to New Excel Workbook
    else:
        print('print pop up window to inform user they cannot put blanks')

# Excel Manipulation Functions
def get_field_label():
    remove_hidden_headers()
    global data_field_list
    sample_list = []
    reject_field_list =['Current Enrolment','Reported Utilization','Waitlist Information (Referral Status)','Not Screened (a)','Screened (b)','On Hold (c)','Waiting time for enrolment','Not Screen', 'Screened - Suitable for Admission']
    data_field_list=[]
    for x in range(1,50): # Cycle through all the rows
        row_data = read_row(ws,x) # Get row_data row by row
        #print("Row",x,":",row_data)
        if x<=6:
            for y in range(0,len(row_data)):
                if y in hidden_list:
                    continue
                else:
                    if row_data[y] != None:
                        if  (isinstance(row_data[y], str)) == True :
                            if row_data[y] in reject_field_list:
                                continue
                            else:
                                sample_list.append(row_data[y])

    for x in range(0,len(FieldLabel_list_DAC)):
            if FieldLabel_list_DAC[x] in sample_list:
                data_field_list.append(FieldLabel_list_DAC[x])

    #if PortFromSheet == 'DAC':
    #    for x in range(0,len(FieldLabel_list_DAC)):
    #        if FieldLabel_list_DAC[x] in sample_list:
    #            data_field_list.append(FieldLabel_list_DAC[x])
    #elif PortFromSheet == 'ADH':
    #    for x in range(0,len(FieldLabel_list_ADH)):
    #        if FieldLabel_list_ADH[x] in sample_list:
    #            data_field_list.append(FieldLabel_list_ADH[x])
    #elif PortFromSheet == 'SW':
    #    for x in range(0,len(FieldLabel_list_SW)):
    #        if FieldLabel_list_SW[x] in sample_list:
    #            data_field_list.append(FieldLabel_list_SW[x])
    #elif PortFromSheet == 'Hostel':
    #    for x in range(0,len(FieldLabel_list_Hostel)):
    #        if FieldLabel_list_Hostel[x] in sample_list:
    #            data_field_list.append(FieldLabel_list_Hostel[x])
    #print(data_field_list)


def remove_hidden_headers():
    global hidden_list
    for colLetter,colDimension in ws.column_dimensions.items():
        if colDimension.hidden == True:
            #print(colDimension.min)
            #print('Last Group',colDimension.max)
            for hidden in range(colDimension.min,colDimension.max+1):
                hidden_list.append(hidden-1)

def remove_hidden_columns(listdata):
    for colLetter,colDimension in ws.column_dimensions.items():
        if colDimension.hidden == True:
            #print(colDimension.min)
            #print('Last Group',colDimension.max)
            for hidden in range(colDimension.min,colDimension.max+1):
                hidden_columns.append(hidden-1)
    data_to_be_sorted = list(listdata.values())
    #print(data_to_be_sorted)
    redict = 1
    for data in data_to_be_sorted:
        count = 0
        for x in hidden_columns:
            data.pop(x-count)
            count+=1
        listdata[redict]=data
        redict+=1
    return(listdata)



def standard_template():
    sheet1 = new_wb["Dec 2021"] #Template Sheet
    sheet2 = new_wb[PortToSheet] # Sheet to copy template to
    maxr = sheet1.max_row
    maxc = sheet1.max_column
    for r in range (1, maxr + 1):
        for c in range (1, maxc + 1):
            if sheet2.cell(row=r,column=c).value == None:
                sheet2.cell(row=r,column=c).value = sheet1.cell(row=r,column=c).value
            else:
                break

def update_data():
    sheet2 = new_wb[PortToSheet] # Sheet to copy template to
    #maxr = sheet2.max_row
    #maxc = sheet2.max_column
    if PortFromSheet == 'SW':
        for r in range (2, len(filtered_Dict) + 2):
            for c in range (1, 5 + 2):
                print((list((list(filtered_Dict.values())[r-2]).values()))[c-1])
                sheet2.cell(row=34+r,column=c).value = (list((list(filtered_Dict.values())[r-2]).values()))[c-1]
    elif PortFromSheet == 'ADH':
        for r in range (2, len(filtered_Dict) + 2):
            for c in range (1, 5 + 2):
                print((list((list(filtered_Dict.values())[r-2]).values()))[c-1])
                sheet2.cell(row=46+r,column=c).value = (list((list(filtered_Dict.values())[r-2]).values()))[c-1]
    elif PortFromSheet == 'Hostel':
        for r in range (2, len(filtered_Dict) + 2):
            for c in range (1, 5 + 2):
                print((list((list(filtered_Dict.values())[r-2]).values()))[c-1])
                sheet2.cell(row=61+r,column=c).value = (list((list(filtered_Dict.values())[r-2]).values()))[c-1]
    else:
        for r in range (2, len(filtered_Dict) + 2):
            for c in range (1, 5 + 2):
                print((list((list(filtered_Dict.values())[r-2]).values()))[c-1])
                sheet2.cell(row=r,column=c).value = (list((list(filtered_Dict.values())[r-2]).values()))[c-1]
    

def read_row(worksheet, row): # Read row cell by cell and return cell values in a list
    return [cell.value for cell in worksheet[row]]

def remove_empty_columns(row_list): # Checks if Cell is None and removes it if it is. Returns a list without None cells.
    emptylist=[]
    for x in row_list: 
        if x is not None:
            emptylist.append(x)
    return(emptylist)

def save_to_list(row_list): # saving data to a list
    emptylist=[]
    for x in row_list:
        emptylist.append(x)
    return(emptylist)

def fill_empty_cell(): # Fill the empty cells that are used in final analysis 
    for x in range(1,len(data_dict)+1):
        for i in range(0,len(data_dict[1])):
            if data_dict[x][i] is None:
                if x <len(data_dict) and x>1: # Check if it's not the first or last row
                    if (isinstance(data_dict[x-1][i], int)) == True or (isinstance(data_dict[x+1][i], int)) == True: # Checking if the upper or lower row cell is filled with a int
                        data_dict[x][i] = 0 # Filling the cell with 0 if the upper or lower cell is filled with a int

                    elif (isinstance(data_dict[x-1][i], str)) == True or (isinstance(data_dict[x+1][i], str)) == True: # Checking if the upper or lower row cell is filled with a str
                        data_dict[x][i] = 'Nil' # Filling the cell with Nil if the upper or lower cell is filled with a str

                elif x==len(data_dict): # Check if it's the last row
                    if (isinstance(data_dict[x-1][i], int)) == True or (isinstance(data_dict[x-2][i], int)) == True: # Checking if the upper or lower row cell is filled with a int
                        data_dict[x][i] = 0 # Filling the cell with 0 if the upper or lower cell is filled with a int

                    elif (isinstance(data_dict[x-1][i], str)) == True or (isinstance(data_dict[x-2][i], str)) == True: # Checking if the upper or lower row cell is filled with a str
                        data_dict[x][i] = 'Nil' # Filling the cell with Nil if the upper or lower cell is filled with a str

                elif x==1 : # Check if it's the first row 
                    if (isinstance(data_dict[x+1][i], int)) == True or (isinstance(data_dict[x+2][i], int)) == True: # Checking if the upper or lower row cell is filled with a int
                        data_dict[x][i] = 0 # Filling the cell with 0 if the upper or lower cell is filled with a int

                    elif (isinstance(data_dict[x+1][i], str)) == True or (isinstance(data_dict[x+2][i], str)) == True: # Checking if the upper or lower row cell is filled with a str
                        data_dict[x][i] = 'Nil' # Filling the cell with Nil if the upper or lower cell is filled with a str

def populate_sheet(worksheet, row,row_data):
    for i in range(1,len(row_data)+1): #loop through all the columns and populate the sheet
            cellref=worksheet.cell(row=row, column=i) #declaring the cell to be given value
            cellref.value=row_data[i-1] #add value to cell

# GUI Display
class UI(QMainWindow):
    def __init__(self):
        super(UI, self).__init__()

        #load the ui file
        uic.loadUi("Main.ui",self)

        # Define our widgets
        self.port_from_button = self.findChild(QPushButton,"Port_From_Button")
        self.port_from_label = self.findChild(QLabel,"Port_From_Label")
        self.port_to_button = self.findChild(QPushButton,"Port_To_Button")
        self.port_to_label = self.findChild(QLabel,"Port_To_Label")
        self.sheet_name_text = self.findChild(QTextEdit,"Sheet_Name_Text")
        self.submit_button = self.findChild(QPushButton,"Submit")
        self.quit_button = self.findChild(QPushButton,"Quit")

        # Click the dropdown box
        self.port_from_button.clicked.connect(self.port_from_clicker)
        self.port_to_button.clicked.connect(self.port_to_clicker)
        self.submit_button.clicked.connect(self.submit_click)
        self.quit_button.clicked.connect(self.quit_click)

        # Show the app
        self.show()

    def quit_click(self):
        self.close()
    
    def submit_click(self):
        global PortToSheet
        PortToSheet = self.sheet_name_text.toPlainText()
        self.close()
        main_loop()
    
    def port_from_clicker(self):
        global PortFromFile, Port_From_Path
        fname = QFileDialog.getOpenFileName(self, "Open Excel File", "","Excel Files (*.xlsx)")

        # Output filename to screen
        if fname:
            Port_From_Path = fname[0]
            print(Port_From_Path)
            x = ((str(fname[0])).split('/'))[-1]
            PortFromFile = x
            self.port_from_label.setText(x)

    def port_to_clicker(self):
        global PortToFile, Port_To_Path
        fname = QFileDialog.getOpenFileName(self, "Open Excel File", "","Excel Files (*.xlsx)")

        # Output filename to screen
        if fname:
            Port_To_Path = fname[0]
            print(Port_To_Path)
            x = ((str(fname[0])).split('/'))[-1]
            PortToFile = x
            self.port_to_label.setText(x)

# Initialize The App
app = QApplication(sys.argv)
UIWindow = UI()
app.exec_()